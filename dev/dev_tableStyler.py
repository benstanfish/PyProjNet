__doc__ = f'Module \"{__name__}\" is used to create base styles for Excel tables and write them to the OOXML stylesheet using Openpyxl'

from userPrefs import settings

import openpyxl as px
from openpyxl.styles import Font, Border, Side, Fill, PatternFill, Color, Alignment
from openpyxl.styles.differential import DifferentialStyle, DifferentialStyleList
from openpyxl.styles.table import TableStyleElement, TableStyle, TableStyleList

from colorfuncs import ContrastText, WebColor

#from openpyxl.writer.theme import theme_xml
# updated_theme_xml = theme_xml.replace('Calibri', 'Arial')
# print(updated_theme_xml)


## TESTING
TEST_FILE_NAME = 'PyTables/tableStylerTests.xlsx'

# wb = px.Workbook()
# ws = wb.active
# wb.save(TEST_FILE_NAME)

wb = px.load_workbook(TEST_FILE_NAME)
sht = wb.active
sht.sheet_view.showGridLines = False


global_font = settings['global']['font_name']
global_font_size = settings['global']['font_size']
wb._fonts[0] = Font(name=global_font, 
                    size=global_font_size)

dxf_whole_table = DifferentialStyle(
    font=Font(name=global_font, size=global_font_size),
    alignment=Alignment(horizontal='left', vertical='top')
)

table_header_backcolor = settings['tables']['colors']['header_background']
table_header_border = settings['tables']['colors']['header_border']

dxf_header_row = DifferentialStyle(
    font=Font(name=global_font, size=global_font_size, color=ContrastText(WebColor(table_header_backcolor)), bold=True),
    fill=PatternFill(bgColor=WebColor(table_header_backcolor)),
    alignment=Alignment(horizontal='left', vertical='bottom'),
    border=Border(bottom=Side(border_style='medium', color=WebColor(table_header_border)))
)

row_line_color = settings['tables']['colors']['row_lines']
dxf_lined_row = DifferentialStyle(
    border=Border(top=Side(style='thin', color=WebColor(row_line_color)))
)

dxf_list = DifferentialStyleList(dxf=[dxf_whole_table,
                                      dxf_header_row,
                                      dxf_lined_row])

whole_table_element = TableStyleElement(type='wholeTable', dxfId=0)
header_element = TableStyleElement(type='headerRow', dxfId=1)
first_row_element = TableStyleElement(type='firstRowStripe', dxfId=2)
second_row_element = TableStyleElement(type='secondRowStripe', dxfId=2)

table_style = TableStyle(name='Base Style', tableStyleElement=[whole_table_element, 
                                                               header_element,
                                                               first_row_element,
                                                               second_row_element])
table_style_list = TableStyleList(tableStyle=[table_style])

wb._differential_styles = dxf_list
wb._table_styles = table_style_list
# wb.loaded_theme = updated_theme_xml










wb.save(TEST_FILE_NAME)

__all__ = []