
from openpyxl import *
from openpyxl.worksheet.table import *

FILE_PATH = 'PyTables/tableRangeTest.xlsx'
FIRST_CELL = 'A5'


wb = Workbook()
# wb = load_workbook(FILE_PATH)
ws = wb.active

header_data = ['ID', 'Status', 'Category', 'Comments', 'Action', 'Assign', 'Date']
# ws.append(header_data)

# for i in range(0, len(header_data)):
#     ws[FIRST_CELL].offset(0, i).value = header_data[i]

# LAST_CELL = ws[FIRST_CELL].offset(0, len(header_data)).coordinate
# header_range = ws[FIRST_CELL:LAST_CELL]      # The range of cells is returned in a list of 1 items

ws.append(header_data)

# for i in range(0, len(header_data)):
#     header_range[0][i].value = header_data[i]

# tab = Table(displayName="fruits", ref="A1:G1")
# style = TableStyleInfo(
#                         name='TableStyleLight9', 
                        # showFirstColumn=False, 
                        # showLastColumn=False,
                        # showRowStripes=True, 
                        # showColumnStripes=True
#                         )
# tab.tableStyleInfo = style
# ws.add_table(tab)

tab = Table(displayName="HelloWorld", ref='A1:G2')
tab.tableStyleInfo = TableStyleInfo(name='TableStyleLight9')
ws.add_table(tab)


# header_range = ws[target, target]
# header_range.value = header_data

wb.save(FILE_PATH)